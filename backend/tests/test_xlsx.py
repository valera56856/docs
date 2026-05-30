"""Behavior tests for the Excel receipt generator (``apps.receipts.services.xlsx``).

The generated ``.xlsx`` is the deliverable a manager imports into SalesDrive
(Склад → Надходження → Імпорт). Its column layout is a contract with that import
template, so these tests pin down:

* The four expected columns, in order:
  ``SKU/Артикул``, ``Назва``, ``Кількість``, ``Ціна (собівартість)``.
* That values come from the *matched product* (SKU + name) and the *line*
  (quantity + cost price), not from the raw OCR fields.
* That :func:`build_receipt_xlsx` returns ``bytes`` that open as a valid
  workbook.

We load the produced bytes back with openpyxl and assert on real cell values, so
a regression in column order, header text, or value sourcing fails loudly.

NOTE: The exact header strings asserted here mirror the manifest. If the real
SalesDrive import template uses different labels, update both the service and
these expectations together.
"""
from __future__ import annotations

import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from apps.catalog.models import OurProduct
from apps.receipts.models import Receipt, ReceiptLine
from apps.receipts.services.xlsx import build_receipt_xlsx
from apps.suppliers.models import Supplier


@pytest.fixture
def receipt_with_lines(db) -> Receipt:
    """Build a receipt with two mapped lines for Excel generation.

    Both lines are mapped to real catalog products and carry an explicit
    quantity and cost price. This is the "ready" shape the generator expects.

    Returns:
        Receipt: A persisted receipt with two ``ReceiptLine`` rows whose
            ``matched_product`` is set.
    """
    supplier = Supplier.objects.create(name="ACME Постачання")
    receipt = Receipt.objects.create(supplier=supplier, status="ready")

    product_a = OurProduct.objects.create(
        salesdrive_id="SD-1", sku="OUR-1", name="Сорочка біла"
    )
    product_b = OurProduct.objects.create(
        salesdrive_id="SD-2", sku="OUR-2", name="Штани чорні"
    )

    ReceiptLine.objects.create(
        receipt=receipt,
        recognized_sku="SUP-A",
        recognized_name="сорочка",
        quantity=Decimal("3.000"),
        price=Decimal("125.50"),
        matched_product=product_a,
        match_status="auto",
    )
    ReceiptLine.objects.create(
        receipt=receipt,
        recognized_sku="SUP-B",
        recognized_name="штани",
        quantity=Decimal("2.000"),
        price=Decimal("899.00"),
        matched_product=product_b,
        match_status="manual",
    )
    return receipt


def _open(xlsx_bytes: bytes):
    """Load workbook bytes and return the active worksheet.

    Args:
        xlsx_bytes: The bytes returned by :func:`build_receipt_xlsx`.

    Returns:
        The active ``openpyxl`` worksheet.
    """
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    return workbook.active


@pytest.mark.django_db
def test_build_receipt_xlsx_returns_bytes(receipt_with_lines: Receipt) -> None:
    """The generator returns non-empty ``bytes`` that open as a workbook.

    Returning raw bytes (rather than writing a file) keeps the service storage
    agnostic — the caller streams them to Cloudflare R2.
    """
    result = build_receipt_xlsx(receipt_with_lines)

    assert isinstance(result, bytes)
    assert len(result) > 0
    # Must open without raising — proves it is a valid xlsx container.
    sheet = _open(result)
    assert sheet is not None


@pytest.mark.django_db
def test_build_receipt_xlsx_has_four_columns_in_order(
    receipt_with_lines: Receipt,
) -> None:
    """Row 1 is exactly the four contract headers, in order.

    Column order is part of the SalesDrive import contract; asserting it here
    catches any accidental reordering or renamed header.
    """
    sheet = _open(build_receipt_xlsx(receipt_with_lines))

    header = [sheet.cell(row=1, column=col).value for col in range(1, 5)]
    assert header == [
        "SKU/Артикул",
        "Назва",
        "Кількість",
        "Ціна (собівартість)",
    ]
    # No stray fifth column header.
    assert sheet.cell(row=1, column=5).value is None


@pytest.mark.django_db
def test_build_receipt_xlsx_values_come_from_matched_product(
    receipt_with_lines: Receipt,
) -> None:
    """Data rows use the *matched product* SKU/name and the *line* qty/price.

    The SKU and name must be the catalog product's values (what SalesDrive
    knows), NOT the recognized supplier SKU/name. The quantity and price are the
    line's editable cost values. We assert the first data row cell-by-cell.
    """
    sheet = _open(build_receipt_xlsx(receipt_with_lines))

    first = receipt_with_lines.lines.order_by("id").first()
    assert first is not None

    sku_cell = sheet.cell(row=2, column=1).value
    name_cell = sheet.cell(row=2, column=2).value
    qty_cell = sheet.cell(row=2, column=3).value
    price_cell = sheet.cell(row=2, column=4).value

    assert sku_cell == first.matched_product.sku  # catalog SKU, not recognized
    assert name_cell == first.matched_product.name
    # Numeric comparison tolerant of Decimal/int/float storage in the cell.
    assert Decimal(str(qty_cell)) == first.quantity
    assert Decimal(str(price_cell)) == first.price


@pytest.mark.django_db
def test_build_receipt_xlsx_writes_one_row_per_line(
    receipt_with_lines: Receipt,
) -> None:
    """Every receipt line produces exactly one data row (plus the header).

    Two lines → header row + two data rows = three populated rows. This guards
    against off-by-one bugs (missing first/last line) in the generator loop.
    """
    sheet = _open(build_receipt_xlsx(receipt_with_lines))

    line_count = receipt_with_lines.lines.count()
    # Count rows that have a SKU in column 1, excluding the header row.
    populated = [
        row
        for row in range(2, 2 + line_count + 5)
        if sheet.cell(row=row, column=1).value not in (None, "")
    ]
    assert len(populated) == line_count


# ---------------------------------------------------------------------------
# Duplicate grouping: same product across multiple lines → one summed row
# ---------------------------------------------------------------------------
def _data_rows(sheet) -> list[list]:
    """Return all non-header data rows (rows with a SKU in column 1).

    Args:
        sheet: An openpyxl worksheet from :func:`_open`.

    Returns:
        A list of ``[sku, name, qty, price]`` row value lists.
    """
    rows: list[list] = []
    for row in range(2, sheet.max_row + 1):
        sku = sheet.cell(row=row, column=1).value
        if sku in (None, ""):
            continue
        rows.append([sheet.cell(row=row, column=col).value for col in range(1, 5)])
    return rows


@pytest.fixture
def receipt_duplicate_lines(db) -> Receipt:
    """Build a receipt where two lines map to the SAME product.

    Line 1: qty 2 @ 100.00; Line 2: qty 3 @ 50.00 — both → ``OUR-1``. A third
    line maps to a different product so we can prove only the duplicates merge.

    Returns:
        Receipt: A persisted receipt with three lines (two sharing a product).
    """
    supplier = Supplier.objects.create(name="ACME Постачання")
    receipt = Receipt.objects.create(supplier=supplier, status="ready")

    product_a = OurProduct.objects.create(
        salesdrive_id="SD-1", sku="OUR-1", name="Сорочка біла"
    )
    product_b = OurProduct.objects.create(
        salesdrive_id="SD-2", sku="OUR-2", name="Штани чорні"
    )

    ReceiptLine.objects.create(
        receipt=receipt,
        recognized_sku="SUP-A1",
        quantity=Decimal("2.000"),
        price=Decimal("100.00"),
        matched_product=product_a,
        match_status="auto",
    )
    ReceiptLine.objects.create(
        receipt=receipt,
        recognized_sku="SUP-A2",
        quantity=Decimal("3.000"),
        price=Decimal("50.00"),
        matched_product=product_a,
        match_status="manual",
    )
    ReceiptLine.objects.create(
        receipt=receipt,
        recognized_sku="SUP-B",
        quantity=Decimal("1.000"),
        price=Decimal("899.00"),
        matched_product=product_b,
        match_status="auto",
    )
    return receipt


@pytest.mark.django_db
def test_duplicate_lines_collapse_to_one_row(
    receipt_duplicate_lines: Receipt,
) -> None:
    """Two lines on the same product produce ONE merged row (plus the other)."""
    sheet = _open(build_receipt_xlsx(receipt_duplicate_lines))
    rows = _data_rows(sheet)

    # Three input lines, but two share a product → two output rows.
    assert len(rows) == 2
    skus = {row[0] for row in rows}
    assert skus == {"OUR-1", "OUR-2"}


@pytest.mark.django_db
def test_duplicate_lines_sum_quantity(receipt_duplicate_lines: Receipt) -> None:
    """The merged row's quantity is the SUM of the duplicates (2 + 3 = 5)."""
    sheet = _open(build_receipt_xlsx(receipt_duplicate_lines))
    merged = next(row for row in _data_rows(sheet) if row[0] == "OUR-1")

    assert Decimal(str(merged[2])) == Decimal("5.000")


@pytest.mark.django_db
def test_duplicate_lines_weighted_average_price(
    receipt_duplicate_lines: Receipt,
) -> None:
    """The merged price is the quantity-weighted average.

    ``Σ(qty·price) / Σ(qty) = (2·100 + 3·50) / (2 + 3) = 350 / 5 = 70.00``.
    This preserves the total cost (350.00) through the merge.
    """
    sheet = _open(build_receipt_xlsx(receipt_duplicate_lines))
    merged = next(row for row in _data_rows(sheet) if row[0] == "OUR-1")

    assert Decimal(str(merged[3])) == Decimal("70.00")
    # Total cost is preserved: 5 units × 70.00 = 350.00 = 2×100 + 3×50.
    assert Decimal(str(merged[2])) * Decimal(str(merged[3])) == Decimal("350.000")
